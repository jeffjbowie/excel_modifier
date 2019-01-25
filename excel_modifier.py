#!/usr/bin/python3
# 3.7.1
# NOTE: Excel is parsed beyond our control in 3rd party app, so we must modify the XLSX files accordingly before processing...

import re
import math
import os
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


# Takes a string input , isolates digits, remove leading zeros.
def round_string(string: str):
    numbers = re.findall(r"[0-9.]+", string)
    for i, number in enumerate(numbers):
        string = string.replace(numbers[i], str(math.ceil(float(number))))
    return string


def capy_sanitize(sheet, row, column):


        first_string = ""
        second_string = ""

        for i in range(7, 13):
            first_string += " " + str(sheet.cell(row=row, column=i).value)
        for i in range(7, 13):
            second_string += " " + str(sheet.cell(row=row+1, column=i).value)

        first_string = first_string.replace("None", "")
        second_string = second_string.replace("None", "")

        # Round all floats , then make sure , is placed in larger #s.

        floats = re.findall(r"([0-9]*\.[0-9]*)", first_string)
        for float_string in floats:
            first_string = first_string.replace(
                float_string,
                  "{:,}".format(int(round(float(float_string)))))

        floats = re.findall(r"([0-9]*\.[0-9]*)", second_string)
        for float_string in floats:
            second_string = second_string.replace(
                float_string,
                 "{:,}".format(int(round(float(float_string)))))

        numbers = re.findall(r"([0-9]{4,})", second_string) 
        for number in numbers:     
             second_string = second_string.replace(str(number), "{:,}".format(int(number)))


        # Blank out 2 rows from cols 7-13

        sheet.cell(row=row, column=7).value = ""
        sheet.cell(row=row, column=8).value = ""
        sheet.cell(row=row, column=9).value = ""
        sheet.cell(row=row, column=10).value = ""
        sheet.cell(row=row, column=11).value = ""
        sheet.cell(row=row, column=12).value = ""
        sheet.cell(row=row, column=13).value = ""

        sheet.cell(row=row+1, column=7).value = ""
        sheet.cell(row=row+1, column=8).value = ""
        sheet.cell(row=row+1, column=9).value = ""
        sheet.cell(row=row+1, column=10).value = ""
        sheet.cell(row=row+1, column=11).value = ""
        sheet.cell(row=row+1, column=12).value = ""
        sheet.cell(row=row+1, column=13).value = ""
        
        sheet.cell(row=row, column=7).value = first_string
        sheet.cell(row=row+1, column=7).value = second_string

# Once we have identified "Additional Stencils" header: 
def addl_stencils_sanitize(sheet, row, column):
    # Check for "X" in the "need" column. 
    if sheet.cell(row=row+3, column=column+1).value == "X":
        # Loop through remaining rows in sheet
        for i in range(row, sheet.max_row+1):
            if sheet.cell(row=i, column=column+6).value is not None and str(sheet.cell(row=i, column=column+6).value).strip() is not "":
                val = str(sheet.cell(row=i, column=column+6).value).strip()
                if "Description" not in val and val is not "" and not val.startswith("QQQQQ"):
                    sheet.cell(row=i, column=column+6).value = "QQQQQ" + val


def process_file(filename=None):

    wb = load_workbook(filename, data_only=True)
    # Iterate over all sheets.
    for sheet in wb:

        if wb.worksheets[0].cell(row=100, column=100).value != "processed":

        # format the Color value as Text
            if wb.worksheets[0].cell(row=7, column=3).value is not None:
                if "WHITE" in wb.worksheets[0].cell(row=7, column=3).value or "BLACK" in wb.worksheets[0].cell(row=7, column=3).value:
                    wb.worksheets[0].cell(row=7, column=3).value = wb.worksheets[0].cell(row=7, column=3).value
                else:
                    wb.worksheets[0].cell(row=7, column=3).value = "YELLOW"
            else:
                wb.worksheets[0].cell(row=7, column=3).value = "YELLOW"

            # Iterate all rows
            for row in sheet.iter_rows():
                # iterate each cell
                for cell in row:
                    if cell.value == "CAPY":
                        capy_sanitize(sheet, cell.row, column_index_from_string(cell.column))
                    elif cell.value == "Additional Stencils":
                        addl_stencils_sanitize(sheet, cell.row, column_index_from_string(cell.column))

   # find the last row & last column , leave a "processed" tag.
    wb.worksheets[0].cell(row=100, column=100).value = "processed"

    wb.save(filename)

# Generate list of files in current directory.
path = os.path.normpath(r"C:\Input")
os.chdir(path)
files = os.listdir(path)

# Loop through & process files.
for file in files:
    if file.endswith('xlsx'):
        process_file(file)